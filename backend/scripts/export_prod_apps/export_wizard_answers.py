"""Export wizard answers from every submitted application into one Excel file.

No file downloads, no auxiliary tables — just the Q&A from the wizard
forms. Output is a single .xlsx with three sheets:

    Index               quick lookup (display id, founder, email, track, status, submitted)
    TIR Applications    one row per submitted TIR app, columns = wizard fields
    SIP Applications    one row per submitted SIP app, columns = wizard fields

Internal columns (id, user_id, created_at, updated_at, completion_pct,
current_section, status) and file-ref JSONB columns are excluded so the
sheet is just the human answers.

Usage:
    cd backend
    source .venv/bin/activate
    pip install openpyxl    # if not already installed

    SUPABASE_URL=<prod-url> \
    SUPABASE_SERVICE_ROLE_KEY=<prod-service-role> \
    OUTPUT_PATH=./wizard_answers.xlsx \
    python scripts/export_prod_apps/export_wizard_answers.py

WARNING: this exports REAL applicant PII (emails, phones, free-text
answers). Treat the output file accordingly.
"""

from __future__ import annotations

import json
import logging
import os
import sys
from typing import Any

try:
    from supabase import create_client
except ImportError:
    print(
        "ERROR: supabase-py not installed.\n"
        "  cd backend && source .venv/bin/activate && pip install -r requirements.txt",
        file=sys.stderr,
    )
    sys.exit(2)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print(
        "ERROR: openpyxl not installed.\n"
        "  pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


# ─── Config ────────────────────────────────────────────────────────────


SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
OUTPUT_PATH = os.environ.get("OUTPUT_PATH", "./wizard_answers.xlsx")
PAGE_SIZE = int(os.environ.get("PAGE_SIZE", "1000"))

if not SUPABASE_URL or not SUPABASE_KEY:
    print(
        "ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set.\n"
        "Set them to the PRODUCTION project's values (not staging).",
        file=sys.stderr,
    )
    sys.exit(2)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("wizard")


# ─── What to exclude from the sheets ───────────────────────────────────


# Internal / non-answer columns. Pinned at the END of the sheet so the
# Q&A is the leftmost content (analysts read these last). Set to empty
# if you want them suppressed entirely.
_TRAILING_COLUMNS: set[str] = {
    "id",
    "user_id",
    "current_section",
    "completion_pct",
    "created_at",
    "updated_at",
}

# Section grouping — column prefix → (sort order, section label). Columns
# without a matching prefix go into the trailing "Other" group.
_SECTION_ORDER: list[tuple[str, str]] = [
    ("display_id",          "Reference"),       # the synthetic id column
    ("track",               "Reference"),
    ("status",              "Reference"),
    ("submitted_at",        "Reference"),
    ("basic_",              "Section 02 · Basic information"),
    ("sip_incorporated",    "Section 02 · Basic information"),
    ("sip_trl",             "Section 02 · Basic information"),
    ("sip_founders",        "Section 02 · Basic information"),
    ("problem_",            "Section 03 · Problem & importance"),
    ("solution_",           "Section 04 · Your solution"),
    ("sip_traction",        "Section 04 · Your solution"),       # SIP traction lives in §04 per the wizard
    ("execution_",          "Section 05 · Execution plan"),
    ("evidence_",           "Section 06 · Evidence"),
    ("sip_pitch_deck",      "Section 06 · Evidence"),
    ("sip_cap_table_file",  "Section 06 · Evidence"),
    ("sip_demo_video_url",  "Section 06 · Evidence"),
    ("sip_patents_files",   "Section 06 · Evidence"),
    ("declaration_",        "Section 07 · Declaration"),
]


def section_for(col: str) -> tuple[int, str]:
    """Return (order_index, label) for a given column. Lower index = leftmost."""
    for idx, (prefix, label) in enumerate(_SECTION_ORDER):
        if col == prefix or (prefix.endswith("_") and col.startswith(prefix)):
            return (idx, label)
    return (len(_SECTION_ORDER), "Other / metadata")


# ─── Friendly column names ─────────────────────────────────────────────


# Pretty headers for known columns. Anything not in here renders the raw
# snake_case name, capitalized.
_LABELS: dict[str, str] = {
    # Reference / metadata
    "display_id":                      "ID",
    "track":                           "Track",
    "status":                          "Status",
    "submitted_at":                    "Submitted at",
    "id":                              "(internal uuid)",
    "user_id":                         "(applicant user uuid)",
    "current_section":                 "(last wizard section)",
    "completion_pct":                  "Completion %",
    "created_at":                      "Created at",
    "updated_at":                      "Updated at",

    # Section 02 · Basic information (TIR legacy + shared)
    "basic_full_name":                 "Full name",
    "basic_phone":                     "Phone",
    "basic_email":                     "Email",
    "basic_org":                       "Organization",
    "basic_degree":                    "Education",
    "basic_has_team":                  "Q · Do you have a team?",
    "basic_teammates":                 "Q · Teammates (JSON)",
    "basic_incubators":                "Q · Incubator(s) (legacy field)",
    "basic_incubator_association":     "Q · Incubator association?",
    "basic_incubator_details":         "Q · Incubator details",
    "basic_hear_about":                "Q · How did you hear about ARTPARK?",
    # Section 02 · SIP-specific
    "sip_incorporated":                "Q · Incorporated as Pvt Ltd?",
    "sip_trl":                         "Q · Technology Readiness Level (TRL)",
    "sip_founders":                    "Q · Founders cap table (JSON)",

    # Section 03 · Problem & importance
    "problem_defined":                 "Q · Have you defined the problem precisely?",
    "problem_describe":                "Q · Describe the problem you are solving",
    "problem_importance":              "Q · Why does this problem matter?",

    # Section 04 · Your solution
    "solution_stage":                  "Q · What stage is your solution at?",
    "solution_describe":               "Q · Describe your solution",
    "solution_core_tech":              "Q · What is the core technology?",
    "solution_ten_x":                  "Q · How is your solution 10× better?",
    "solution_hurdles":                "Q · Hurdles to building this",
    "solution_moat":                   "Q · What is your moat?",
    "solution_national_scale":         "Q · How does this scale nationally?",
    "solution_customers":              "Q · Who are your customers?",
    "solution_contrarian_insight":     "Q · Contrarian insight",
    # SIP traction (lives in section 04 per the wizard)
    "sip_traction":                    "Q · Current traction status",
    "sip_traction_details":            "Q · Traction details",
    "sip_traction_files":              "Q · Traction supporting files (paths)",

    # Section 05 · Execution plan
    "execution_will_break":            "Q · What will break in execution?",
    "execution_milestone":             "Q · Next 12-month milestone",
    "execution_milestone_files":       "Q · Milestone supporting files (paths)",
    "execution_budget":                "Q · Budget breakdown",
    "execution_failure":               "Q · How could the execution fail?",
    "execution_infrastructure":        "Q · Infrastructure / resource needs",
    "execution_hwsw_integration":      "Q · Hardware-software integration approach",

    # Section 06 · Evidence (TIR legacy + SIP)
    "evidence_files":                  "Evidence files (paths)",
    "evidence_video_url":              "Demo video URL (TIR)",
    "evidence_deck":                   "Pitch deck (TIR)",
    "sip_pitch_deck":                  "Pitch deck (SIP, JSON)",
    "sip_cap_table_file":              "Cap table file (SIP, JSON)",
    "sip_demo_video_url":              "Demo video URL (SIP)",
    "sip_patents_files":               "Patents / publications (paths)",

    # Section 07 · Declaration
    "declaration_truthful":            "✓ Truthful declaration",
    "declaration_ref_checks":          "✓ Reference checks allowed",
    "declaration_terms":               "✓ Terms accepted",
    "declaration_newsletter":          "✓ Newsletter opt-in",
}


def label_for(col: str) -> str:
    return _LABELS.get(col, col.replace("_", " ").capitalize())


# ─── Display ID (same shape as the leadership dashboard) ──────────────


def compose_display_id(track: str, app_id: str | None) -> str:
    prefix = (track or "?").upper()
    if not app_id:
        return f"{prefix}-?????"
    try:
        n = int(str(app_id).replace("-", "")[:8], 16) % 100_000
        return f"{prefix}-{n:05d}"
    except (ValueError, TypeError):
        return f"{prefix}-{str(app_id)[:5]}"


# ─── Fetch helpers ─────────────────────────────────────────────────────


def paginate(client, table: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    page = 0
    while True:
        start = page * PAGE_SIZE
        end = start + PAGE_SIZE - 1
        try:
            res = client.table(table).select("*").range(start, end).execute()
            rows = res.data or []
        except Exception as exc:
            log.warning("paginate(%s) page=%d failed: %s", table, page, exc)
            break
        out.extend(rows)
        if len(rows) < PAGE_SIZE:
            break
        page += 1
    return out


def paginate_with_fallback(client, primary: str, fallback: str | None) -> tuple[list[dict[str, Any]], str | None]:
    """Try `primary` first; if it returns 0 rows (likely because the table
    doesn't exist in the schema cache), retry against `fallback`. Returns
    (rows, table_used). table_used is None when both fail.

    This handles the prod-vs-staging schema split: migration 010 renamed
    `applications` → `tir_applications` and added `sip_applications`. Prod
    may still be on the legacy schema.
    """
    rows = paginate(client, primary)
    if rows:
        return rows, primary
    if not fallback:
        return rows, primary if rows else None
    log.info("  %s returned 0 rows — trying legacy table %s", primary, fallback)
    rows = paginate(client, fallback)
    return rows, (fallback if rows else None)


def cell_value(v: Any) -> Any:
    """Convert a column value into something Excel can store.

    JSONB dicts/lists are pretty-printed JSON strings; booleans become
    'Yes'/'No' for readability; everything else passes through.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (dict, list)):
        return json.dumps(v, indent=2, ensure_ascii=False)
    return v


# ─── Sheet writers ─────────────────────────────────────────────────────


def _column_order(rows: list[dict[str, Any]]) -> list[str]:
    """Order every column from the row union by wizard section, falling
    back to insertion order within a section. Trailing metadata columns
    (id, user_id, timestamps) land at the very end."""
    all_cols: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                all_cols.append(k)
    # Make sure the synthesised reference columns are in the set even
    # if the row dict didn't carry them yet.
    for k in ("display_id", "track"):
        if k not in seen:
            all_cols.append(k)
            seen.add(k)

    def sort_key(col: str) -> tuple[int, int]:
        # Trailing cols → very large index so they end up rightmost.
        trail_idx = 999 if col in _TRAILING_COLUMNS else section_for(col)[0]
        return (trail_idx, all_cols.index(col))

    return sorted(all_cols, key=sort_key)


def _write_track_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet(name)
    if not rows:
        ws.cell(row=1, column=1, value="(no submitted applications)")
        return 0

    cols = _column_order(rows)

    # ── Row 1: section header (merged across each section's columns) ──
    # Walk the columns left-to-right, group consecutive runs of the same
    # section label, merge their cells in row 1.
    run_start_ci = 1
    run_label = section_for(cols[0])[1] if cols[0] not in _TRAILING_COLUMNS else "Reference / metadata"
    for ci, col in enumerate(cols, start=1):
        label = ("Reference / metadata" if col in _TRAILING_COLUMNS
                 else section_for(col)[1])
        if ci == 1:
            run_label = label
            continue
        if label != run_label:
            if ci - 1 > run_start_ci:
                ws.merge_cells(start_row=1, start_column=run_start_ci,
                               end_row=1, end_column=ci - 1)
            hcell = ws.cell(row=1, column=run_start_ci, value=run_label)
            hcell.font = Font(bold=True, size=11, color="FFFFFF")
            hcell.fill = PatternFill("solid", fgColor="3213B7")  # ARTBlue
            hcell.alignment = Alignment(horizontal="center", vertical="center")
            run_start_ci = ci
            run_label = label
    # close the final run
    if len(cols) > run_start_ci:
        ws.merge_cells(start_row=1, start_column=run_start_ci,
                       end_row=1, end_column=len(cols))
    hcell = ws.cell(row=1, column=run_start_ci, value=run_label)
    hcell.font = Font(bold=True, size=11, color="FFFFFF")
    hcell.fill = PatternFill("solid", fgColor="3213B7")
    hcell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: question labels ────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        c = ws.cell(row=2, column=ci, value=label_for(col))
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="EFEFEF")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Rows 3+: data ─────────────────────────────────────────────────
    for ri, row in enumerate(rows, start=3):
        for ci, col in enumerate(cols, start=1):
            v = cell_value(row.get(col))
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Freeze the two header rows + first 3 columns (id/track/status).
    ws.freeze_panes = "D3"

    # Column widths.
    narrow = {"display_id", "track", "status", "submitted_at",
              "basic_degree", "sip_incorporated", "sip_trl",
              "completion_pct", "declaration_truthful", "declaration_ref_checks",
              "declaration_terms", "declaration_newsletter"}
    medium = {"basic_full_name", "basic_email", "basic_org", "basic_phone",
              "basic_hear_about", "basic_incubator_association",
              "basic_has_team", "id", "user_id"}
    for ci, col in enumerate(cols, start=1):
        letter = ws.cell(row=2, column=ci).column_letter
        if col in narrow:
            ws.column_dimensions[letter].width = 16
        elif col in medium:
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 60

    # Tall enough for wrap on the question-label row.
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 40
    # Tall enough for wrap on the answer rows.
    for ri in range(3, len(rows) + 3):
        ws.row_dimensions[ri].height = 60

    return len(rows)


def _write_index_sheet(wb: Workbook, all_rows: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("Index", 0)  # first sheet
    headers = ["ID", "Track", "Status", "Submitted at",
               "Founder", "Email", "Organization"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor="EFEFEF")

    for ri, r in enumerate(all_rows, start=2):
        ws.cell(row=ri, column=1, value=r.get("display_id"))
        ws.cell(row=ri, column=2, value=(r.get("track") or "").upper())
        ws.cell(row=ri, column=3, value=r.get("status"))
        ws.cell(row=ri, column=4, value=r.get("submitted_at"))
        ws.cell(row=ri, column=5, value=r.get("basic_full_name"))
        ws.cell(row=ri, column=6, value=r.get("basic_email"))
        ws.cell(row=ri, column=7, value=r.get("basic_org"))

    ws.freeze_panes = "A2"
    widths = [12, 8, 14, 24, 28, 32, 28]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    return len(all_rows)


# ─── Main ──────────────────────────────────────────────────────────────


def main() -> int:
    log.info("supabase: %s", SUPABASE_URL)
    log.info("output:   %s", OUTPUT_PATH)

    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    log.info("fetching tir_applications (with legacy `applications` fallback)…")
    tir_all, tir_table = paginate_with_fallback(
        client, "tir_applications", fallback="applications",
    )
    log.info("  %d total rows (from %s)", len(tir_all), tir_table or "(none)")
    tir = [r for r in tir_all if r.get("status") and r.get("status") != "draft"]
    log.info("  %d non-draft (will be exported)", len(tir))
    for r in tir:
        r["track"] = "tir"
        r["display_id"] = compose_display_id("tir", r.get("id"))

    log.info("fetching sip_applications…")
    sip_all, _ = paginate_with_fallback(client, "sip_applications", fallback=None)
    log.info("  %d total rows", len(sip_all))
    sip = [r for r in sip_all if r.get("status") and r.get("status") != "draft"]
    log.info("  %d non-draft (will be exported)", len(sip))
    for r in sip:
        r["track"] = "sip"
        r["display_id"] = compose_display_id("sip", r.get("id"))

    # Combined index — sorted newest-first by submitted_at
    combined = sorted(
        tir + sip,
        key=lambda r: r.get("submitted_at") or r.get("created_at") or "",
        reverse=True,
    )

    log.info("building workbook…")
    wb = Workbook()
    # Drop the default blank sheet
    default = wb.active
    wb.remove(default)

    n_index = _write_index_sheet(wb, combined)
    n_tir = _write_track_sheet(wb, "TIR Applications", tir)
    n_sip = _write_track_sheet(wb, "SIP Applications", sip)

    output_path = os.path.abspath(OUTPUT_PATH)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    log.info("wrote %s: %d index rows, %d TIR, %d SIP",
             output_path, n_index, n_tir, n_sip)
    return 0


if __name__ == "__main__":
    sys.exit(main())
