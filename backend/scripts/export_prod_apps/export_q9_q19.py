"""Export Q9–Q19 answers for every submitted TIR application since a cutoff
date, with the template's exact question text as column headers.

Maps to migration 008's canonical Q→column table:

    Q9  → problem_describe              Q15 → execution_will_break
    Q10 → problem_defined               Q16 → execution_milestone
    Q11 → solution_describe             Q17 → execution_infrastructure
    Q12 → solution_core_tech            Q18 → execution_failure
    Q13 → solution_contrarian_insight   Q19 → execution_hwsw_integration
    Q14 → solution_stage

Q15 has a fallback: legacy `solution_hurdles` is read if
`execution_will_break` is NULL (the wizard renamed this slot once during
2026; older submissions may live in the legacy column).

Usage:
    cd backend
    source .venv/bin/activate
    pip install openpyxl python-docx     # one-time

    SUPABASE_URL=https://xtmszlpwgbyoumalgbhs.supabase.co \\
    SUPABASE_SERVICE_ROLE_KEY=<prod-key> \\
    CUTOFF_DATE=2026-04-28 \\
    OUTPUT_PATH="$HOME/Desktop/tir_q9_q19_responses.xlsx" \\
    python scripts/export_prod_apps/export_q9_q19.py

Outputs both `<OUTPUT_PATH>` (.xlsx) and a sibling `.csv` next to it.
"""

from __future__ import annotations

import csv
import logging
import os
import sys
from typing import Any

try:
    from supabase import create_client
except ImportError:
    print("ERROR: supabase-py not installed", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed — run `pip install openpyxl`", file=sys.stderr)
    sys.exit(2)


SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
CUTOFF_DATE  = os.environ.get("CUTOFF_DATE", "2026-04-28").strip()
OUTPUT_PATH  = os.environ.get("OUTPUT_PATH",
                              os.path.expanduser("~/Desktop/tir_q9_q19_responses.xlsx"))
PAGE_SIZE    = int(os.environ.get("PAGE_SIZE", "1000"))

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY required", file=sys.stderr)
    sys.exit(2)

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("q9_q19")


# ─── Template question text (verbatim from TIR_Application_Response_Template.docx) ──
# Headers shown in the export. Keep the wording in lockstep with the docx
# so an applicant + reviewer reading both files see the same question.
QUESTIONS: list[tuple[str, str, str, str]] = [
    # (Q-id, primary db column, fallback db column, question title)
    ("Q9",  "problem_describe",            "",
        "Q9 · What specific \"critical problem\" in your chosen sector are you solving?"),
    ("Q10", "problem_defined",             "",
        "Q10 · Do you think the problem you want to solve is well-defined?"),
    ("Q11", "solution_describe",           "",
        "Q11 · Describe your solution."),
    ("Q12", "solution_core_tech",          "",
        "Q12 · What's the core technology that makes this special and hard to replicate?"),
    ("Q13", "solution_contrarian_insight", "",
        "Q13 · Share a genuinely rare insight in your field of expertise."),
    ("Q14", "solution_stage",              "",
        "Q14 · How far along are you?"),
    ("Q15", "execution_will_break",        "solution_hurdles",
        "Q15 · What are the primary technical hurdles you need to overcome?"),
    ("Q16", "execution_milestone",         "",
        "Q16 · What are the most critical milestone(s) you aim to achieve during this residency?"),
    ("Q17", "execution_infrastructure",    "",
        "Q17 · What specific advanced infrastructure or facilities are essential for your success during this residency?"),
    ("Q18", "execution_failure",           "",
        "Q18 · Tell us about a significant research direction or prototype failure — how did you pivot, and what did it teach you?"),
    ("Q19", "execution_hwsw_integration",  "",
        "Q19 · How do you manage complex hardware-software integration?"),
]


# Basic identity columns shown to the left of the Q-columns. The recipient
# wants to know whose answers they're reading.
IDENTITY_COLUMNS: list[tuple[str, str]] = [
    ("display_id",    "ID"),
    ("submitted_at",  "Submitted at"),
    ("status",        "Status"),
    ("basic_full_name", "Founder name"),
    ("basic_email",   "Email"),
    ("basic_phone",   "Phone"),
    ("basic_org",     "Organization"),
]


# ─── Helpers ───────────────────────────────────────────────────────────


def compose_display_id(app_id: str | None) -> str:
    if not app_id:
        return "TIR-?????"
    try:
        n = int(str(app_id).replace("-", "")[:8], 16) % 100_000
        return f"TIR-{n:05d}"
    except (ValueError, TypeError):
        return f"TIR-{str(app_id)[:5]}"


def paginate(client, table: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    page = 0
    while True:
        start = page * PAGE_SIZE
        end = start + PAGE_SIZE - 1
        try:
            res = client.table(table).select("*").range(start, end).execute()
            rows = res.data or []
        except Exception as exc:
            log.warning("paginate(%s) page=%d failed: %s", table, page, exc)
            break
        out.extend(rows)
        if len(rows) < PAGE_SIZE:
            break
        page += 1
    return out


def cell(row: dict[str, Any], primary: str, fallback: str = "") -> Any:
    """Read primary; fall back if primary is None/empty."""
    v = row.get(primary)
    if v in (None, "", [], {}) and fallback:
        v = row.get(fallback)
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (list, dict)):
        # Empty containers → "—"; otherwise stringify.
        if not v:
            return ""
        return repr(v)
    return v if v is not None else ""


# ─── Main ──────────────────────────────────────────────────────────────


def main() -> int:
    log.info("supabase: %s", SUPABASE_URL)
    log.info("cutoff:   %s", CUTOFF_DATE)
    log.info("output:   %s", OUTPUT_PATH)

    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    log.info("fetching applications (legacy table)…")
    rows = paginate(client, "applications")
    log.info("  %d total rows", len(rows))

    submitted = [
        r for r in rows
        if r.get("status") and r.get("status") != "draft"
        and (r.get("submitted_at") or "") >= CUTOFF_DATE
    ]
    log.info("  %d submitted on or after %s (will be exported)", len(submitted), CUTOFF_DATE)

    # Decorate with display_id; sort by submitted_at desc
    for r in submitted:
        r["display_id"] = compose_display_id(r.get("id"))
    submitted.sort(key=lambda r: r.get("submitted_at") or "", reverse=True)

    # ── Build the column list ────────────────────────────────────────
    header_cells: list[str] = [label for _, label in IDENTITY_COLUMNS]
    header_cells.extend([title for _, _, _, title in QUESTIONS])
    keys: list[tuple[str, str]] = [(col, "") for col, _ in IDENTITY_COLUMNS]
    keys.extend([(prim, fb) for _, prim, fb, _ in QUESTIONS])

    # ── CSV output ───────────────────────────────────────────────────
    csv_path = os.path.splitext(OUTPUT_PATH)[0] + ".csv"
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header_cells)
        for r in submitted:
            w.writerow([cell(r, primary, fallback) for primary, fallback in keys])
    log.info("wrote %s (%d rows)", csv_path, len(submitted))

    # ── Excel output ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Q9–Q19 responses"

    # Header row
    for ci, h in enumerate(header_cells, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3213B7")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Data rows
    for ri, row in enumerate(submitted, start=2):
        for ci, (primary, fb) in enumerate(keys, start=1):
            c = ws.cell(row=ri, column=ci, value=cell(row, primary, fb))
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    id_widths = {0: 14, 1: 22, 2: 12, 3: 24, 4: 28, 5: 16, 6: 28}
    for ci in range(1, len(header_cells) + 1):
        letter = ws.cell(row=1, column=ci).column_letter
        if ci - 1 in id_widths:
            ws.column_dimensions[letter].width = id_widths[ci - 1]
        else:
            ws.column_dimensions[letter].width = 60  # Q-columns

    ws.freeze_panes = "B2"  # freeze ID column + header row
    ws.row_dimensions[1].height = 60
    for ri in range(2, len(submitted) + 2):
        ws.row_dimensions[ri].height = 80  # generous wrap for long answers

    os.makedirs(os.path.dirname(OUTPUT_PATH) or ".", exist_ok=True)
    wb.save(OUTPUT_PATH)
    log.info("wrote %s (%d rows)", OUTPUT_PATH, len(submitted))
    log.info("done")
    return 0


if __name__ == "__main__":
    sys.exit(main())
