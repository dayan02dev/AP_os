"""mis_export: turns one or more period bundles (mis_query.period_bundle's
own shape, plus admin_vip_query.fetch_mis_period's `application_id`/
`startup` overlay) into a catalog-column-ordered report -- xlsx via
openpyxl, csv via the stdlib csv module. Pure/service-level: no DB, no
FastAPI. Column order and derived-not-stored values are exercised here;
the /admin/platform/vip/mis/export endpoint (test_admin_vip_mis_export.py)
covers the HTTP wiring on top.
"""
from __future__ import annotations

import csv
import io

import openpyxl
import pytest

from app.services import mis_export


def _monthly_bundle(startup="Acme Robotics", application_id="app1"):
    return {
        "application_id": application_id,
        "startup": startup,
        "period": {
            "period_key": "2026-08", "label": "Aug 2026", "kind": "monthly",
            "status": "submitted", "due_date": "2026-09-05", "submitted_at": "2026-08-20T00:00:00Z",
        },
        "metrics": [
            {"metric_key": "revenue_month", "label": "Revenue this month (₹ Lakh)",
             "group_key": "commercial", "unit": "₹L", "target": 10, "actual": 12.5, "commentary": "ahead"},
            {"metric_key": "trl_level", "label": "TRL Level (1–9)",
             "group_key": "product_technology", "unit": "1–9", "target": None, "actual": 6, "commentary": None},
        ],
        "financials": [],
        "headcount": [],
        "entries": {
            "milestones": [
                {"id": "e1", "period_id": "p1", "section": "milestones", "sort_order": 0,
                 "data": {"milestone": "Tape out v0.3", "owner": "Priya", "status": "On Track", "notes": ""}},
            ],
            "risks": [], "asks": [],
        },
        "narrative": {"exec.headline_win": "Shipped v2", "exec.cash": "4.2 Cr in bank"},
        "derived": {
            "metrics": {"vs_last": {"revenue_month": 2.5, "trl_level": None}},
            "financials": {"needs_gap": {}},
            "headcount": {"net_change": {}, "total": {"current_count": None, "exited": None}},
        },
    }


def _quarterly_bundle(startup="Acme Robotics", application_id="app1"):
    return {
        "application_id": application_id,
        "startup": startup,
        "period": {
            "period_key": "FY26-27-Q1", "label": "Q1 FY26-27", "kind": "quarterly",
            "status": "submitted", "due_date": "2026-07-15", "submitted_at": "2026-07-10T00:00:00Z",
        },
        "metrics": [],
        "financials": [
            {"series": "needs_total", "bucket": "Q1 (Current)", "amount": 100},
            {"series": "needs_confirmed", "bucket": "Q1 (Current)", "amount": 40},
            {"series": "needs_projected", "bucket": "Q1 (Current)", "amount": 30},
        ],
        "headcount": [
            {"category": "artpark_associated", "current_count": 12, "exited": 1, "remarks": None},
            {"category": "startup", "current_count": 8, "exited": 0, "remarks": None},
            {"category": "consultants", "current_count": 2, "exited": 0, "remarks": None},
            {"category": "interns", "current_count": 1, "exited": 0, "remarks": None},
        ],
        "entries": {
            "ip_assets": [], "collaborations": [], "publications": [], "products": [],
            "funding": [], "planned_vs_actual": [], "next_milestones": [],
        },
        "narrative": {"glance.strategic_theme": "First commercial deployments"},
        "derived": {
            "metrics": {"vs_last": {}},
            "financials": {"needs_gap": {"Q1 (Current)": 30, "Q2 (Next)": None, "Q3": None, "Q4": None, "Q5": None}},
            "headcount": {
                "net_change": {"artpark_associated": -3, "startup": 2, "consultants": 0, "interns": 1},
                "total": {"current_count": 23, "exited": 1},
            },
        },
    }


# ── validation: no fail-open defaults ─────────────────────────────────────

def test_unknown_kind_raises():
    with pytest.raises(ValueError):
        mis_export.build_sheets(kind="weekly", scope="startup", bundles=[_monthly_bundle()])


def test_unknown_scope_raises():
    with pytest.raises(ValueError):
        mis_export.build_sheets(kind="monthly", scope="cohort-ish", bundles=[_monthly_bundle()])


# ── column order comes from the catalog ───────────────────────────────────

def test_metrics_sheet_column_order_matches_catalog():
    sheets = mis_export.build_sheets(kind="monthly", scope="startup", bundles=[_monthly_bundle()])
    metrics_sheet = next(s for s in sheets if s.name == "Key Metrics")
    assert metrics_sheet.headers == ["Metric", "Group", "Unit", "Target", "Actual", "vs Last Mo", "Commentary"]
    row = next(r for r in metrics_sheet.rows if r[0] == "Revenue this month (₹ Lakh)")
    assert row == ["Revenue this month (₹ Lakh)", "Commercial", "₹L", 10, 12.5, 2.5, "ahead"]


def test_entries_sheet_column_order_matches_entry_fields():
    sheets = mis_export.build_sheets(kind="monthly", scope="startup", bundles=[_monthly_bundle()])
    milestones = next(s for s in sheets if s.name == "Milestones")
    assert milestones.headers == ["Milestone", "Owner", "Status", "Notes"]
    assert milestones.rows == [["Tape out v0.3", "Priya", "On Track", ""]]


def test_narrative_sheet_lists_only_fields_the_period_actually_has():
    sheets = mis_export.build_sheets(kind="monthly", scope="startup", bundles=[_monthly_bundle()])
    narrative = next(s for s in sheets if s.name == "Narrative")
    values = {r[2]: r[3] for r in narrative.rows}  # (prompt col index 2, value col index 3)
    assert values.get("Headline win") == "Shipped v2"
    assert "Biggest concern" not in values  # never answered in this bundle


# ── cohort scope: one row per startup per section ─────────────────────────

def test_cohort_scope_produces_one_row_per_startup():
    bundles = [_monthly_bundle(startup="Acme Robotics", application_id="app1"),
               _monthly_bundle(startup="Beta Labs", application_id="app2")]
    sheets = mis_export.build_sheets(kind="monthly", scope="cohort", bundles=bundles)
    metrics_sheet = next(s for s in sheets if s.name == "Key Metrics")
    assert metrics_sheet.headers[0] == "Startup"
    revenue_rows = [r for r in metrics_sheet.rows if r[1] == "Revenue this month (₹ Lakh)"]
    assert {r[0] for r in revenue_rows} == {"Acme Robotics", "Beta Labs"}


def test_startup_scope_has_no_startup_column():
    sheets = mis_export.build_sheets(kind="monthly", scope="startup", bundles=[_monthly_bundle()])
    metrics_sheet = next(s for s in sheets if s.name == "Key Metrics")
    assert metrics_sheet.headers[0] == "Metric"


# ── derived, never stored: computed fresh for the report, not copied ──────

def test_financials_sheet_shows_computed_gap_not_a_stored_column():
    sheets = mis_export.build_sheets(kind="quarterly", scope="startup", bundles=[_quarterly_bundle()])
    financials = next(s for s in sheets if s.name == "Financials")
    gap_row = next(r for r in financials.rows if r[0] == "Gap" and r[1] == "Q1 (Current)")
    assert gap_row[2] == 30  # 100 - 40 - 30, from derived.financials.needs_gap -- never a "needs_gap" input row


def test_headcount_sheet_total_row_has_blank_net_change():
    """Mirrors the source template exactly: the Total row carries Current
    Count / Exited only -- Net Change is left blank, never summed."""
    sheets = mis_export.build_sheets(kind="quarterly", scope="startup", bundles=[_quarterly_bundle()])
    headcount = next(s for s in sheets if s.name == "Headcount")
    total_row = next(r for r in headcount.rows if r[0] == "Total")
    assert total_row[1] == 23  # current_count
    assert total_row[2] == 1   # exited
    assert total_row[3] is None  # net_change -- deliberately blank


def test_headcount_sheet_net_change_is_the_derived_value_not_current_minus_exited():
    sheets = mis_export.build_sheets(kind="quarterly", scope="startup", bundles=[_quarterly_bundle()])
    headcount = next(s for s in sheets if s.name == "Headcount")
    row = next(r for r in headcount.rows if r[0] == "Employees (ARTPARK, associated with startup)")
    # current_count=12, exited=1 -> current-exited would wrongly read 11;
    # the true (derived) net_change supplied is -3.
    assert row[3] == -3


# ── xlsx rendering ─────────────────────────────────────────────────────────

def test_render_xlsx_round_trips_through_openpyxl():
    sheets = mis_export.build_sheets(kind="monthly", scope="startup", bundles=[_monthly_bundle()])
    data = mis_export.render_xlsx(sheets)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Key Metrics" in wb.sheetnames
    ws = wb["Key Metrics"]
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == ["Metric", "Group", "Unit", "Target", "Actual", "vs Last Mo", "Commentary"]


def test_render_xlsx_sheet_names_are_valid_and_unique():
    sheets = mis_export.build_sheets(kind="quarterly", scope="startup", bundles=[_quarterly_bundle()])
    data = mis_export.render_xlsx(sheets)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert len(wb.sheetnames) == len({n.lower() for n in wb.sheetnames})
    assert all(len(n) <= 31 for n in wb.sheetnames)


# ── csv rendering ───────────────────────────────────────────────────────

def test_render_csv_contains_a_block_per_sheet():
    sheets = mis_export.build_sheets(kind="monthly", scope="startup", bundles=[_monthly_bundle()])
    text = mis_export.render_csv(sheets)
    assert "## Key Metrics" in text
    assert "## Milestones" in text
    reader = csv.reader(io.StringIO(text.split("## Key Metrics\n", 1)[1]))
    header = next(reader)
    assert header == ["Metric", "Group", "Unit", "Target", "Actual", "vs Last Mo", "Commentary"]
