"""GET /admin/platform/vip/mis/export: capability gate, startup vs cohort
scope, xlsx/csv rendering, wired against real founder-submitted state (not
hand-inserted rows) the same way test_admin_vip.py builds its fixtures.
"""
from __future__ import annotations

import io

import openpyxl
import pytest
from app.deps import get_current_user
from app.main import app

from tests.fixtures.fake_supabase import FakeSupabase

CUR_MONTH = "2026-08"


@pytest.fixture
def _clear():
    yield
    app.dependency_overrides.clear()


@pytest.fixture(autouse=True)
def _frozen_today(freezer):
    freezer.move_to("2026-08-16T06:00:00Z")


class _Bucket:
    def upload(self, *a, **k): return {"path": a[0] if a else ""}
    def create_signed_url(self, path, expires_in): return {"signedURL": f"https://x/{path}"}


class _Storage:
    def from_(self, bucket): return _Bucket()


def _founder_user(app_id: str = "sapp1"):
    return lambda: {"user_id": "u1", "email": "u1@x.com", "roles": ["applicant"]}


def _admin_user(role: str = "admin"):
    return lambda: {"user_id": "admin1", "email": "admin1@x.com", "roles": [role]}


def _as(user_fn):
    app.dependency_overrides[get_current_user] = user_fn


def _install(monkeypatch, extra_sip_apps: list[dict] | None = None):
    from app.routers import admin_vip_mis_export as export_router  # noqa: F401
    from app.routers import founder as founder_router
    from app.routers import founder_mis as mis_router
    from app.services import admin_vip_query, air_query, applications_query, mis_query

    sip_apps = [
        {"id": "sapp1", "user_id": "u1", "status": "onboarded",
         "submitted_at": "2026-01-01", "basic_org": "Acme Robotics"},
    ] + (extra_sip_apps or [])

    tables = {
        "sip_applications": sip_apps,
        "tir_applications": [],
        "application_status_log": [
            {"id": f"log-{a['id']}", "application_id": a["id"], "application_track": "sip",
             "from_status": "offered", "to_status": "onboarded",
             "changed_at": "2026-06-01T00:00:00+00:00"}
            for a in sip_apps
        ],
        "ai_screening": [],
        "vip_air_assessments": [], "vip_air_lever_scores": [],
        "vip_mis_periods": [], "vip_mis_metrics": [], "vip_mis_financials": [],
        "vip_mis_headcount": [], "vip_mis_entries": [],
    }
    fake = FakeSupabase(tables)
    fake.storage = _Storage()
    for mod in (founder_router, mis_router, air_query, mis_query, admin_vip_query, applications_query):
        monkeypatch.setattr(mod, "get_admin_client", lambda f=fake: f)
    return fake


def test_export_requires_view_all_apps_capability(client, monkeypatch, _clear):
    _install(monkeypatch)
    _as(_founder_user())
    r = client.get(f"/admin/platform/vip/mis/export?kind=monthly&period={CUR_MONTH}&application_id=sapp1")
    assert r.status_code == 403
    assert r.json()["detail"]["code"] == "missing_capability"


def test_export_startup_scope_requires_application_id(client, monkeypatch, _clear):
    _install(monkeypatch)
    _as(_admin_user())
    r = client.get(f"/admin/platform/vip/mis/export?kind=monthly&period={CUR_MONTH}")
    assert r.status_code == 422
    assert r.json()["detail"]["code"] == "application_id_required"


def test_export_unknown_kind_is_422(client, monkeypatch, _clear):
    _install(monkeypatch)
    _as(_admin_user())
    r = client.get(f"/admin/platform/vip/mis/export?kind=weekly&period={CUR_MONTH}&application_id=sapp1")
    assert r.status_code == 422


def test_export_startup_scope_xlsx(client, monkeypatch, _clear):
    _install(monkeypatch)
    _as(_founder_user())
    client.get("/founder/mis")  # generates the period rows
    client.put(f"/founder/mis/monthly/{CUR_MONTH}/metrics", json=[
        {"metric_key": "revenue_month", "actual": 12.5, "target": 10},
    ])

    _as(_admin_user())
    r = client.get(
        f"/admin/platform/vip/mis/export?kind=monthly&period={CUR_MONTH}"
        "&scope=startup&format=xlsx&application_id=sapp1"
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in r.headers["content-disposition"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Key Metrics" in wb.sheetnames
    ws = wb["Key Metrics"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert header[0] == "Metric"  # no leading Startup column for scope=startup
    revenue_row = next(r for r in rows[1:] if r[0] == "Revenue this month (₹ Lakh)")
    assert revenue_row[header.index("Actual")] == 12.5


def test_export_cohort_scope_includes_startup_column_and_only_startups_with_the_period(client, monkeypatch, _clear):
    fake = _install(monkeypatch, extra_sip_apps=[
        {"id": "sapp2", "user_id": "u2", "status": "onboarded",
         "submitted_at": "2026-01-01", "basic_org": "Beta Labs"},
    ])
    _as(_founder_user("sapp1"))
    client.get("/founder/mis")

    _as(lambda: {"user_id": "u2", "email": "u2@x.com", "roles": ["applicant"]})
    client.get("/founder/mis")

    _as(_admin_user())
    r = client.get(
        f"/admin/platform/vip/mis/export?kind=monthly&period={CUR_MONTH}&scope=cohort&format=xlsx"
    )
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Key Metrics"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert header[0] == "Startup"
    startups = {r[0] for r in rows[1:]}
    assert startups == {"Acme Robotics", "Beta Labs"}


def test_export_csv_format(client, monkeypatch, _clear):
    _install(monkeypatch)
    _as(_founder_user())
    client.get("/founder/mis")

    _as(_admin_user())
    r = client.get(
        f"/admin/platform/vip/mis/export?kind=monthly&period={CUR_MONTH}"
        "&scope=startup&format=csv&application_id=sapp1"
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    assert "## Key Metrics" in r.text


def test_export_leadership_role_also_allowed(client, monkeypatch, _clear):
    _install(monkeypatch)
    _as(_founder_user())
    client.get("/founder/mis")
    _as(_admin_user(role="leadership"))
    r = client.get(
        f"/admin/platform/vip/mis/export?kind=monthly&period={CUR_MONTH}&application_id=sapp1"
    )
    assert r.status_code == 200, r.text
